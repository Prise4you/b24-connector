"""
Bitrix24 → Knowledge Base connector v2.
Полный extract задач со всеми полями, комментариями, ссылками.

Использование:
  python connector.py --dry-run              # собрать документы, не загружать
  python connector.py                        # собрать + сохранить в out/
  python connector.py --skip-admin-check     # без проверки прав (для тестов)
  python connector.py --poll                 # инкрементальный запуск по изменениям

С удалённым управлением (через б24-admin на хостинге):
  python connector.py \
    --config-url https://anit.ru/b24-admin/api/config.php \
    --status-url https://anit.ru/b24-admin/api/status.php \
    --connector-token TOKEN

  Или задать CONNECTOR_TOKEN=... как переменную окружения (для launchd).
"""
import json
import os
import re
import sys
import time
import argparse
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from b24_client import B24Client, B24Error
import security

# ─── конфиг ───────────────────────────────────────────────────────────────────

def load_config(path: str = "config.json") -> dict:
    # Если путь относительный — ищем рядом со скриптом
    p = Path(path)
    if not p.is_absolute() and not p.exists():
        p = Path(__file__).parent / path
    with open(p, encoding="utf-8") as f:
        return json.load(f)


def fetch_remote_config(url: str, token: str) -> dict:
    """Скачать config.json с хостинга (GET, X-Connector-Token)."""
    req = urllib.request.Request(
        url + "?raw=1",
        headers={"X-Connector-Token": token},
    )
    with urllib.request.urlopen(req, timeout=15) as resp:
        data = json.loads(resp.read().decode())
    # Сохраняем локально как резервную копию
    local = Path(__file__).parent / "config_remote_cache.json"
    local.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    os.chmod(local, 0o600)
    return data


def post_run_status(url: str, token: str, payload: dict) -> None:
    """Отправить результат прогона на хостинг (POST, X-Connector-Token)."""
    body = json.dumps(payload, ensure_ascii=False).encode()
    req = urllib.request.Request(
        url,
        data=body,
        headers={"Content-Type": "application/json", "X-Connector-Token": token},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            resp.read()
    except Exception as e:
        print(f"  ⚠️  Не удалось отправить статус на хостинг: {e}")


def push_remote_config(url: str, token: str, cfg: dict) -> None:
    """
    Отправить обновлённый config.json обратно на хостинг (POST, тот же
    api/config.php, что использует и панель — принимает X-Connector-Token).
    Нужно вызывать сразу после изменения notebook_id при работе через
    --config-url: без этого id создаваемого ноутбука существует только в
    памяти эфемерного прогона (напр. GitHub Actions раннера) и теряется
    навсегда, вызывая создание нового ноутбука при каждом следующем прогоне.
    """
    body = json.dumps(cfg, ensure_ascii=False).encode()
    req = urllib.request.Request(
        url,
        data=body,
        headers={"Content-Type": "application/json", "X-Connector-Token": token},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=15) as resp:
        resp.read()


def push_notebook_id_patch(url: str, token: str, kind: str, notebook_id: str,
                            group_id: int = None) -> None:
    """
    Точечно обновить notebook_id ОДНОГО проекта (или CRM) на хостинге, не
    перезаписывая весь config.json. Нужно для matrix-параллелизации
    (несколько job'ов coннектора для разных проектов работают одновременно —
    полная перезапись всего конфига одним из них могла бы затереть
    notebook_id, только что сохранённый другим).
    """
    patch = {"kind": kind, "notebook_id": notebook_id}
    if group_id is not None:
        patch["group_id"] = group_id
    body = json.dumps({"patch": patch}, ensure_ascii=False).encode()
    req = urllib.request.Request(
        url, data=body,
        headers={"Content-Type": "application/json", "X-Connector-Token": token},
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=15) as resp:
        resp.read()


def save_config(cfg: dict, args, connector_token: str) -> None:
    """
    Сохранить изменённый конфиг там, откуда он был загружен: на хостинг
    (если работали через --config-url) или в локальный файл.
    """
    if args.config_url and connector_token:
        try:
            push_remote_config(args.config_url, connector_token, cfg)
            print("  Конфиг обновлён на хостинге")
        except Exception as e:
            print(f"  ⚠️  Не удалось сохранить конфиг на хостинге ({e})")
    else:
        with open(args.config, "w", encoding="utf-8") as fh:
            json.dump(cfg, fh, ensure_ascii=False, indent=2)
        security.harden_config_file(args.config)

# ─── справочники ──────────────────────────────────────────────────────────────

PRIORITY_LABELS = {"0": "Низкий", "1": "Средний", "2": "Высокий"}
STATUS_LABELS = {
    "1": "Ждёт выполнения", "2": "В работе", "3": "Завершена (ожидание)",
    "4": "Отложена",        "5": "Завершена", "6": "Почти просрочена",
    "7": "Просрочена"
}

# ─── обработка текста ─────────────────────────────────────────────────────────

def _extract_urls(text: str) -> list:
    """Вытащить все URL из BBcode [URL=...] и обычных http-ссылок."""
    urls = re.findall(r'\[URL=([^\]]+)\]', text)
    urls += re.findall(r'(?<!\[URL=)https?://\S+', text)
    return list(dict.fromkeys(u.rstrip(']') for u in urls))

_DISK_CACHE = {}  # fid → строка описания (кэш в пределах прогона)

def _resolve_disk_refs(text: str, c: B24Client) -> tuple:
    """
    Найти [DISK FILE ID=xxx] в тексте.
    Резолвить через disk.file.get с кэшем (один ID не запрашивается дважды).
    Вернуть (очищенный текст, список строк с описанием файлов).
    """
    file_notes = []
    ids = re.findall(r'\[DISK FILE ID=n?(\d+)\]', text)
    for fid in ids:
        if fid in _DISK_CACHE:
            file_notes.append(_DISK_CACHE[fid])
            continue
        try:
            f = c.call("disk.file.get", {"id": int(fid)})
            name = f.get("NAME", f"файл #{fid}")
            ext  = f.get("EXTENSION", "")
            size = f.get("SIZE", "")
            note = f"📎 {name} ({ext.upper()}, {size} байт) [ID={fid}]"
        except B24Error:
            note = f"📎 файл #{fid} (удалён или недоступен)"
        _DISK_CACHE[fid] = note
        file_notes.append(note)
    clean = re.sub(r'\[DISK FILE ID=n?\d+\]', '', text)
    return clean, file_notes

def clean_bbcode(text: str, preserve_users: bool = True) -> str:
    """Очистить BBcode, сохранив имена из [USER=xxx]Name[/USER]."""
    if not text:
        return ""
    if preserve_users:
        text = re.sub(r'\[USER=\d+\]([^\[]+)\[/USER\]', r'@\1', text)
    text = re.sub(r'\[URL=([^\]]+)\]([^\[]*)\[/URL\]', r'\2 (\1)', text)
    text = re.sub(r'\[/?[A-Z][A-Z0-9]*[^\]]*\]', '', text)
    return text.strip()

def fmt_date(iso: str) -> str:
    return iso[:10] if iso else ""

def fmt_datetime(iso: str) -> str:
    return iso[:16].replace("T", " ") if iso else ""

# ─── person helpers ───────────────────────────────────────────────────────────

def person_label(obj: dict) -> str:
    """Из embedded объекта {name, workPosition} → строка."""
    if not obj:
        return ""
    name = obj.get("name", "")
    pos  = obj.get("workPosition", "")
    return f"{name} ({pos})" if pos else name

def auditors_label(data) -> str:
    """auditorsData: dict или list → строка имён."""
    if not data:
        return ""
    if isinstance(data, dict):
        items = data.values()
    else:
        items = data
    return ", ".join(person_label(u) for u in items if u)

def accomplices_label(data) -> str:
    if not data:
        return ""
    if isinstance(data, dict):
        items = data.values()
    else:
        items = data
    return ", ".join(person_label(u) for u in items if u)


def _format_multifield(data) -> str:
    """Форматировать CRM-поля phone/email: [{VALUE, VALUE_TYPE}...] → строка."""
    if not data:
        return ""
    if isinstance(data, list):
        return ", ".join(item.get("VALUE", "") for item in data if item.get("VALUE"))
    return str(data)


def _fmt_msg_files(msg: dict) -> list:
    """Извлечь метаданные файлов из сообщения IM."""
    files = msg.get("files") or msg.get("ATTACH") or []
    if isinstance(files, dict):
        files = list(files.values())
    lines = []
    for f in files:
        if not isinstance(f, dict):
            continue
        f_type = (f.get("type") or f.get("TYPE") or "FILE").upper()
        f_name = f.get("name") or f.get("NAME") or "файл"
        f_size = f.get("size") or f.get("SIZE") or ""
        f_url  = f.get("urlDownload") or f.get("link") or ""
        size_str = f" ({f_size} байт)" if f_size else ""
        if f_type == "AUDIO":
            line = f"  🎵 Аудио: {f_name}{size_str}"
        elif f_type == "IMAGE":
            line = f"  🖼 Изображение: {f_name}{size_str}"
        elif f_type == "VIDEO":
            line = f"  🎬 Видео: {f_name}{size_str}"
        else:
            line = f"  📎 Файл: {f_name}{size_str}"
        if f_url:
            line += f" → {str(f_url)[:200]}"
        lines.append(line)
    return lines

_AUDIO_EXTS = {".ogg", ".mp3", ".m4a", ".wav", ".opus", ".flac", ".aiff", ".aac"}


def _download_bytes(url: str) -> bytes:
    """Скачать байты по URL. Возвращает пустые байты при ошибке или HTML-ответе."""
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "b24-connector/1.0"})
        with urllib.request.urlopen(req, timeout=60) as resp:
            ct = resp.headers.get("Content-Type", "")
            if "text/html" in ct:
                return b""
            return resp.read()
    except Exception:
        return b""


def download_audio_from_messages(messages: list, audio_dir: str) -> list:
    """
    Скачать аудиофайлы из сообщений чата в audio_dir.
    Возвращает [(dest_path, title), ...] для загрузки в NotebookLM.
    """
    result = []
    for msg in messages:
        files = msg.get("files") or msg.get("ATTACH") or []
        if isinstance(files, dict):
            files = list(files.values())
        for f in files:
            if not isinstance(f, dict):
                continue
            f_type  = (f.get("type") or f.get("TYPE") or "").upper()
            ext     = ("." + f.get("extension", "").lower()) if f.get("extension") else ""
            f_name  = f.get("name") or f.get("NAME") or f"audio{ext or '.ogg'}"
            url     = f.get("urlDownload") or f.get("link") or ""
            if f_type != "AUDIO" and ext not in _AUDIO_EXTS:
                continue
            if not url:
                continue
            safe_name = re.sub(r'[^\w.\-]', '_', f_name)
            dest = str(Path(audio_dir) / safe_name)
            data = _download_bytes(url)
            if data:
                Path(audio_dir).mkdir(parents=True, exist_ok=True)
                Path(dest).write_bytes(data)
                kb = len(data) // 1024
                print(f"    🎵 Скачано: {f_name} ({kb} KB)")
                result.append((dest, f"Аудио: {f_name}"))
            else:
                print(f"    ⚠️  Не удалось скачать: {f_name}")
    return result


# ─── EXTRACT ──────────────────────────────────────────────────────────────────

def _chunks(seq, n):
    for i in range(0, len(seq), n):
        yield seq[i:i + n]

def extract_tasks_full(c: B24Client, group_id: int) -> list:
    """
    Загрузить все задачи группы с полным набором полей через tasks.task.get.
    Индивидуальные вызовы (не batch) — надёжнее при больших проектах.
    select=* возвращает embedded creator/responsible/auditorsData/accomplicesData/checkListTree.
    """
    print(f"  Загружаю список задач группы {group_id}...")
    ids = [t["id"] for t in c.call_list("tasks.task.list", {
        "filter": {"GROUP_ID": group_id},
        "select": ["ID"],
    }, result_key="tasks")]
    print(f"    Найдено задач: {len(ids)}")

    tasks = []
    for i, tid in enumerate(ids, 1):
        try:
            item = c.call("tasks.task.get", {"taskId": tid, "select": ["*"]})
            if item and isinstance(item, dict) and "task" in item:
                tasks.append(item["task"])
        except B24Error as e:
            print(f"    ⚠️  Задача {tid}: {e} — пропускаю")
        if i % 20 == 0 or i == len(ids):
            print(f"    Загружено: {len(tasks)}/{len(ids)}")
    return tasks

def extract_comments_bulk(c: B24Client, task_ids: list) -> dict:
    """Пакетно загрузить комментарии для всех задач. Возвращает {task_id: [comments]}."""
    print(f"  Загружаю комментарии ({len(task_ids)} задач, batch)...")
    result = {}
    failed_chunks = 0
    for chunk in _chunks(task_ids, 15):
        cmds = {f"c{tid}": ("task.commentitem.getlist",
                            {"TASKID": tid, "ORDER": {"ID": "ASC"}})
                for tid in chunk}
        res = None
        for attempt in range(2):
            try:
                res = c.batch(cmds)
                break
            except B24Error as e:
                if attempt == 0:
                    print(f"    ⚠️  Чанк комментариев не загружен ({e}), повтор...")
                    time.sleep(2.0)
                else:
                    print(f"    ❌ Чанк комментариев пропущен после повтора: {e}")
                    failed_chunks += 1
        if res is None:
            for tid in chunk:
                result.setdefault(str(tid), [])
            continue
        inner = res.get("result", {})
        for tid in chunk:
            item = inner.get(f"c{tid}")
            result[str(tid)] = item if isinstance(item, list) else []
    total = sum(len(v) for v in result.values())
    print(f"    Комментариев всего: {total}" +
          (f" (пропущено чанков: {failed_chunks})" if failed_chunks else ""))
    return result

def extract_group_chat(c: B24Client, group_id: int) -> tuple:
    try:
        dialog = c.call("im.dialog.get", {"DIALOG_ID": f"SG{group_id}"})
    except B24Error as e:
        print(f"    Чат недоступен: {e}")
        return None, []
    if not dialog:
        return None, []

    print(f"  Загружаю сообщения чата (dialog {dialog.get('id')})...")
    messages = []
    try:
        for msg in c.call_list("im.dialog.messages.get", {
            "DIALOG_ID": f"SG{group_id}", "LIMIT": 50
        }, result_key="messages"):
            messages.append(msg)
    except B24Error as e:
        print(f"    Ошибка сообщений: {e}")
    print(f"    Сообщений: {len(messages)}")
    return dialog, messages

def extract_disk_files(c: B24Client, group_id: int) -> list:
    print(f"  Загружаю файлы диска группы {group_id}...")
    try:
        storages = c.call("disk.storage.getlist", {
            "filter": {"ENTITY_TYPE": "group", "ENTITY_ID": group_id}
        })
    except B24Error as e:
        print(f"    Диск недоступен: {e}")
        return []
    if not storages:
        return []
    root_id = storages[0].get("ROOT_OBJECT_ID")
    if not root_id:
        return []
    files = []
    _walk_folder(c, int(root_id), files, depth=0)
    print(f"    Файлов: {len(files)}")
    return files

def _walk_folder(c, folder_id, files, depth):
    if depth > 4:
        return
    try:
        children = c.call("disk.folder.getchildren", {"id": folder_id})
    except B24Error:
        return
    if not isinstance(children, list):
        return
    for item in children:
        if item.get("TYPE") == "folder":
            _walk_folder(c, int(item["ID"]), files, depth + 1)
        else:
            files.append(item)


# ─── содержимое файлов диска (не только метаданные) ───────────────────────────

_ASIS_EXTS = {"pdf", "png", "jpg", "jpeg", "gif", "webp"}
_TEXT_CONVERTIBLE_EXTS = {"docx", "xlsx"}
_INLINE_TEXT_EXTS = {"txt", "md", "csv"}
_MAX_DISK_FILE_BYTES = 20 * 1024 * 1024
_MAX_INLINE_TEXT_BYTES = 100 * 1024


def _disk_file_download_url(c: B24Client, file_id) -> str:
    try:
        f = c.call("disk.file.get", {"id": int(file_id)})
        return f.get("DOWNLOAD_URL", "")
    except B24Error:
        return ""


def convert_docx_to_text(data: bytes) -> str:
    import io
    from docx import Document
    doc = Document(io.BytesIO(data))
    return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())


def convert_xlsx_to_markdown(data: bytes, max_rows: int = 200) -> str:
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    lines = []
    for ws in wb.worksheets:
        lines.append(f"## Лист «{ws.title}»")
        row_count = 0
        for row in ws.iter_rows(values_only=True):
            if row_count >= max_rows:
                lines.append(f"_...усечено, показаны первые {max_rows} строк_")
                break
            cells = [str(v) if v is not None else "" for v in row]
            if any(cells):
                lines.append("| " + " | ".join(cells) + " |")
            row_count += 1
        lines.append("")
    return "\n".join(lines)


def prepare_disk_attachments(c: B24Client, files: list, group_name: str,
                             proj_dir: str, limit: int = 15) -> tuple:
    """
    Скачать содержимое top-N файлов диска (по дате изменения) и подготовить
    к загрузке в NotebookLM:
      - PDF/изображения — as-is, отдельными бинарными source (NotebookLM
        парсит PDF нативно; конвертация не нужна). Сознательно БЕЗ прогона
        через security.redact_doc — бинарные файлы физически не читаемы
        текстовым фильтром секретов (см. connector/CLAUDE.md).
      - Word (.docx)/Excel (.xlsx) — конвертация в текст, уходит в общий
        docs{} и, значит, через обычный redact_doc наравне с tasks.md и т.д.
      - Мелкие текстовые файлы (.txt/.md/.csv) — встраиваются прямо в
        files.md, не как отдельный source (экономия лимита источников).
    Остальные (вне top-N/неподдерживаемые) — только метаданные, как раньше.

    Возвращает (text_docs, binary_attachments, inline_sections):
      text_docs — {"disk_<file_id>.md": content} для docs{} (пройдёт redact)
      binary_attachments — [(path, title)] для upload_doc as-is
      inline_sections — [str] дополнительные секции для files.md
    """
    text_docs = {}
    binary_attachments = []
    inline_sections = []

    real_files = [f for f in files if f.get("TYPE") != "folder"]
    real_files.sort(key=lambda f: f.get("UPDATE_TIME") or f.get("CREATE_TIME") or "",
                     reverse=True)
    top = real_files[:limit]

    attach_dir = Path(proj_dir) / "disk_attachments"

    for f in top:
        name = f.get("NAME", "")
        ext = (f.get("EXTENSION") or Path(name).suffix.lstrip(".")).lower()
        file_id = f.get("ID")
        size = int(f.get("SIZE") or 0)
        if not file_id or size > _MAX_DISK_FILE_BYTES:
            continue

        if ext in _INLINE_TEXT_EXTS:
            if size > _MAX_INLINE_TEXT_BYTES:
                continue
            url = _disk_file_download_url(c, file_id)
            data = _download_bytes(url) if url else b""
            if not data:
                continue
            try:
                text = data.decode("utf-8")
            except UnicodeDecodeError:
                text = data.decode("cp1251", errors="replace")
            inline_sections.append(f"\n### {name}\n\n{text[:5000]}\n")
            continue

        if ext not in _ASIS_EXTS and ext not in _TEXT_CONVERTIBLE_EXTS:
            continue

        url = _disk_file_download_url(c, file_id)
        data = _download_bytes(url) if url else b""
        if not data:
            continue

        if ext in _ASIS_EXTS:
            attach_dir.mkdir(parents=True, exist_ok=True)
            safe_name = re.sub(r'[^\w.\-]', '_', name)
            dest = attach_dir / f"{file_id}_{safe_name}"
            dest.write_bytes(data)
            binary_attachments.append((str(dest), f"{group_name} — Файл: {name}"))
            print(f"    📎 as-is: {name} ({size // 1024} KB)")
        else:
            try:
                if ext == "docx":
                    text = convert_docx_to_text(data)
                else:
                    text = convert_xlsx_to_markdown(data)
            except Exception as e:
                print(f"    ⚠️  не удалось сконвертировать {name}: {e}")
                continue
            if text.strip():
                text_docs[f"disk_{file_id}.md"] = (
                    f"# {group_name} — {name}\n\n{text}"
                )
                print(f"    📄 сконвертирован: {name} ({len(text)} симв.)")

    return text_docs, binary_attachments, inline_sections

# ─── EXTRACT: CRM ─────────────────────────────────────────────────────────────

def extract_crm_companies(c: B24Client, ids: list) -> list:
    print("  Загружаю компании CRM...")
    select = ["ID", "TITLE", "INDUSTRY", "PHONE", "EMAIL",
              "WEB", "ADDRESS", "COMMENTS", "DATE_CREATE", "DATE_MODIFY",
              "ASSIGNED_BY_ID", "SOURCE_ID"]
    if ids:
        companies = []
        for company_id in ids:
            try:
                raw = c.call("crm.company.get", {"id": company_id, "select": select})
            except Exception as e:
                print(f"    ⚠️  CRM компания {company_id}: {e} — пропущена")
                continue
            if raw:
                companies.append(raw)
    else:
        companies = list(c.call_list("crm.company.list", {"select": select}))
    print(f"    Компаний: {len(companies)}")
    return companies


def extract_crm_contacts(c: B24Client, ids: list) -> list:
    print("  Загружаю контакты CRM...")
    select = ["ID", "NAME", "LAST_NAME", "SECOND_NAME",
              "PHONE", "EMAIL", "COMPANY_ID", "COMPANY_TITLE",
              "POST", "SOURCE_ID", "DATE_CREATE", "DATE_MODIFY",
              "ASSIGNED_BY_ID", "COMMENTS"]
    if ids:
        contacts = []
        for contact_id in ids:
            try:
                raw = c.call("crm.contact.get", {"id": contact_id, "select": select})
            except Exception as e:
                print(f"    ⚠️  CRM контакт {contact_id}: {e} — пропущен")
                continue
            if raw:
                contacts.append(raw)
    else:
        contacts = list(c.call_list("crm.contact.list", {"select": select}))
    print(f"    Контактов: {len(contacts)}")
    return contacts


def extract_crm_deals(c: B24Client, company_ids: list = None) -> list:
    print("  Загружаю сделки CRM...")
    params = {
        "select": ["ID", "TITLE", "STAGE_ID", "OPPORTUNITY", "CURRENCY_ID",
                   "COMPANY_ID", "CONTACT_ID", "ASSIGNED_BY_ID",
                   "CLOSEDATE", "DATE_CREATE", "COMMENTS"],
        "filter": {"ACTIVE": "Y"},
    }
    if company_ids:
        params["filter"]["COMPANY_ID"] = company_ids
    deals = list(c.call_list("crm.deal.list", params))
    print(f"    Сделок: {len(deals)}")
    return deals


def extract_crm_contacts_by_company(c: B24Client, company_ids: list) -> list:
    """Контакты, привязанные к указанным компаниям (фильтр COMPANY_ID)."""
    if not company_ids:
        return []
    print("  Загружаю контакты компаний CRM...")
    select = ["ID", "NAME", "LAST_NAME", "SECOND_NAME",
              "PHONE", "EMAIL", "COMPANY_ID", "COMPANY_TITLE",
              "POST", "SOURCE_ID", "DATE_CREATE", "DATE_MODIFY",
              "ASSIGNED_BY_ID", "COMMENTS"]
    contacts = list(c.call_list("crm.contact.list", {
        "select": select,
        "filter": {"COMPANY_ID": company_ids},
    }))
    print(f"    Контактов: {len(contacts)}")
    return contacts


def extract_crm_smart(c: B24Client, company_ids: list = None) -> list:
    print("  Загружаю смарт-процессы CRM...")
    try:
        raw = c.call("crm.type.list", {})
        types = raw.get("types") or [] if isinstance(raw, dict) else []
    except Exception:
        return []
    want = set(str(x) for x in (company_ids or []))
    items = []
    for t in types:
        entity_type_id = t.get("entityTypeId")
        if not entity_type_id:
            continue
        try:
            params = {
                "entityTypeId": entity_type_id,
                "select": ["id", "title", "stageId", "assignedById",
                           "companyId", "createdTime", "updatedTime"],
            }
            if want:
                params["filter"] = {"companyId": list(want)}
            rows = c.call_list("crm.item.list", params)
            # подстраховка: если фильтр по companyId не отработал на стороне Б24,
            # отсеиваем в Python (оставляем только элементы с нужной компанией)
            if want:
                rows = [r for r in rows if str(r.get("companyId") or "") in want]
            for r in rows:
                r["_typeName"] = t.get("title", f"Тип {entity_type_id}")
            items.extend(rows)
        except Exception:
            pass
    print(f"    Смарт-элементов: {len(items)}")
    return items


_TIMELINE_MAX_PER_ENTITY = 100  # cap на карточку — не тянуть всю историю старой сделки


def extract_crm_timeline(c: B24Client, entities: list) -> list:
    """
    История комментариев (crm.timeline.comment.list) по списку сущностей
    entities: [{"type": "company"|"contact"|"deal", "id": ..., "name": ...}].
    Ограничено _TIMELINE_MAX_PER_ENTITY на карточку — не вся история, а
    последние записи (сортировка по CREATED, свежие первыми).
    """
    print("  Загружаю историю комментариев CRM (timeline)...")
    result = []
    for ent in entities:
        entity_type, entity_id = ent.get("type"), ent.get("id")
        if not entity_type or not entity_id:
            continue
        try:
            rows = list(c.call_list("crm.timeline.comment.list", {
                "filter": {"ENTITY_ID": entity_id, "ENTITY_TYPE": entity_type},
                "select": ["ID", "CREATED", "AUTHOR_ID", "COMMENT"],
            }))
        except B24Error:
            continue
        rows.sort(key=lambda r: r.get("CREATED") or "", reverse=True)
        for r in rows[:_TIMELINE_MAX_PER_ENTITY]:
            r["_entity_name"] = ent.get("name", f"{entity_type} #{entity_id}")
        result.extend(rows[:_TIMELINE_MAX_PER_ENTITY])
    print(f"    Комментариев (timeline): {len(result)}")
    return result


def extract_crm_activities(c: B24Client, entities: list) -> list:
    """
    Завершённые активности (звонки/письма/встречи, БЕЗ записи звонка —
    только метаданные: тема, направление, дата) через crm.activity.list.
    entities: [{"type": "company"|"contact"|"deal", "id": ..., "name": ...}].
    """
    print("  Загружаю активности CRM...")
    _OWNER_TYPE_ID = {"company": 4, "contact": 3, "deal": 2}
    result = []
    for ent in entities:
        entity_type, entity_id = ent.get("type"), ent.get("id")
        owner_type_id = _OWNER_TYPE_ID.get(entity_type)
        if not owner_type_id or not entity_id:
            continue
        try:
            rows = list(c.call_list("crm.activity.list", {
                "filter": {"OWNER_TYPE_ID": owner_type_id, "OWNER_ID": entity_id,
                           "COMPLETED": "Y"},
                "select": ["ID", "SUBJECT", "DESCRIPTION", "DIRECTION",
                           "TYPE_ID", "CREATED", "END_TIME"],
            }))
        except B24Error:
            continue
        rows.sort(key=lambda r: r.get("CREATED") or "", reverse=True)
        for r in rows[:_TIMELINE_MAX_PER_ENTITY]:
            r["_entity_name"] = ent.get("name", f"{entity_type} #{entity_id}")
        result.extend(rows[:_TIMELINE_MAX_PER_ENTITY])
    print(f"    Активностей: {len(result)}")
    return result


# ─── TRANSFORM: задачи ────────────────────────────────────────────────────────

def build_tasks_doc(tasks: list, group_name: str, c: B24Client,
                    comments_by_task: dict) -> str:
    # Кэш id→title для резолва parentId
    task_titles = {t["id"]: t.get("title", f"#{t['id']}") for t in tasks}

    lines = [
        f"# {group_name} — Задачи и обсуждения",
        f"_Сгенерировано: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}_",
        f"_Всего задач: {len(tasks)}_",
        ""
    ]

    for t in tasks:
        tid     = t.get("id", "")
        title   = t.get("title", "")
        status  = STATUS_LABELS.get(str(t.get("status", "")), t.get("status", ""))
        priority = PRIORITY_LABELS.get(str(t.get("priority", "1")), "Средний")

        creator     = person_label(t.get("creator") or {})
        responsible = person_label(t.get("responsible") or {})
        auditors    = auditors_label(t.get("auditorsData") or t.get("auditors") or [])
        accomplices = accomplices_label(t.get("accomplicesData") or t.get("accomplices") or [])

        created  = fmt_date(t.get("createdDate", ""))
        deadline = fmt_date(t.get("deadline", ""))
        closed   = fmt_datetime(t.get("closedDate", ""))
        changed  = fmt_datetime(t.get("changedDate", ""))

        # Закрыл
        closed_by_id = t.get("closedBy", "")
        closed_by = ""
        if closed_by_id:
            # Имя закрывшего: если совпадает с creator/responsible — берём оттуда
            for role_obj in [t.get("creator"), t.get("responsible")]:
                if role_obj and str(role_obj.get("id", "")) == str(closed_by_id):
                    closed_by = role_obj.get("name", str(closed_by_id))
                    break
            if not closed_by:
                # Ищем в наблюдателях
                aud = t.get("auditorsData") or {}
                if isinstance(aud, dict) and str(closed_by_id) in aud:
                    closed_by = aud[str(closed_by_id)].get("name", str(closed_by_id))
                else:
                    closed_by = str(closed_by_id)

        parent_id = t.get("parentId")
        parent_str = f"подзадача → «{task_titles.get(str(parent_id), f'#{parent_id}')}»" if parent_id else ""

        task_control = "да" if t.get("taskControl") == "Y" else ""
        crm_link = t.get("ufCrmTask") or ""

        # Теги
        tags_raw = t.get("tags") or []
        if isinstance(tags_raw, dict):
            tags_raw = list(tags_raw.values())
        tags = ", ".join(str(tg.get("title", tg)) if isinstance(tg, dict) else str(tg)
                         for tg in tags_raw)

        # Описание: извлечь URL и файловые ссылки
        raw_desc = t.get("description", "") or ""
        urls = _extract_urls(raw_desc)
        raw_desc, file_refs = _resolve_disk_refs(raw_desc, c)
        desc = clean_bbcode(raw_desc)

        # Чеклист
        checklist = []
        tree = t.get("checkListTree") or {}
        for node in (tree.get("descendants") or []):
            f = node.get("fields", {})
            if f.get("title"):
                mark = "✅" if f.get("isComplete") else "⬜"
                checklist.append(f"{mark} {f['title']}")

        # Прикреплённые файлы
        attached_file_ids = t.get("ufTaskWebdavFiles") or []

        # ─── формируем секцию задачи ─────────────────────────────────────────
        lines.append(f"## Задача #{tid}: {title}")

        meta = []
        meta.append(f"**Статус:** {status}")
        meta.append(f"**Приоритет:** {priority}")
        if creator:
            meta.append(f"**Постановщик:** {creator}")
        if responsible:
            meta.append(f"**Ответственный:** {responsible}")
        if accomplices:
            meta.append(f"**Соисполнители:** {accomplices}")
        if auditors:
            meta.append(f"**Наблюдатели:** {auditors}")
        if created:
            meta.append(f"**Создана:** {created}")
        if deadline:
            meta.append(f"**Дедлайн:** {deadline}")
        if closed:
            closed_info = f"**Закрыта:** {closed}"
            if closed_by:
                closed_info += f" ({closed_by})"
            meta.append(closed_info)
        if changed:
            meta.append(f"**Изменена:** {changed}")
        if parent_str:
            meta.append(f"**{parent_str}**")
        if task_control:
            meta.append(f"**Требует проверки:** {task_control}")
        if crm_link:
            meta.append(f"**CRM:** {crm_link}")
        if tags:
            meta.append(f"**Теги:** {tags}")
        lines.append("  ".join(meta))

        if desc:
            lines.append("")
            lines.append(desc[:2000])

        if urls:
            lines.append("")
            lines.append("**Ссылки в описании:**")
            for u in urls[:10]:
                lines.append(f"- {u}")

        if file_refs:
            lines.append("")
            for fr in file_refs:
                lines.append(fr)

        if attached_file_ids:
            lines.append("")
            lines.append(f"**Прикреплённые файлы (IDs):** {', '.join(str(i) for i in attached_file_ids)}")

        if checklist:
            lines.append("")
            lines.append("**Чеклист:**")
            for item in checklist:
                lines.append(f"  {item}")

        if comments_by_task is not None:
            comments = comments_by_task.get(str(tid), [])
            if comments:
                lines.append("")
                lines.append("**Комментарии:**")
                for cm in comments:
                    author = cm.get("AUTHOR_NAME", str(cm.get("AUTHOR_ID", "")))
                    dt     = fmt_datetime(cm.get("POST_DATE", ""))
                    raw_cm = cm.get("POST_MESSAGE", "") or ""
                    cm_urls = _extract_urls(raw_cm)
                    raw_cm, cm_files = _resolve_disk_refs(raw_cm, c)
                    text = clean_bbcode(raw_cm)
                    if text:
                        lines.append(f"- [{dt}] **{author}:** {text[:800]}")
                    for u in cm_urls[:5]:
                        lines.append(f"  → {u}")
                    for fr in cm_files:
                        lines.append(f"  {fr}")

        lines.append("")
        lines.append("---")
        lines.append("")

    return "\n".join(lines)

# ─── TRANSFORM: чат ───────────────────────────────────────────────────────────

def build_chat_doc(dialog: dict, messages: list, group_name: str) -> str:
    name = (dialog.get("name", group_name) if dialog else group_name)
    lines = [
        f"# {group_name} — Чат группы",
        f"_Сгенерировано: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}_",
        f"_Название: {name}_",
        f"_Сообщений в выгрузке: {len(messages)}_",
        ""
    ]
    for msg in reversed(messages):
        author = msg.get("authorId", "")
        date   = fmt_date(msg.get("date") or msg.get("DATE_CREATE", ""))
        text   = clean_bbcode(msg.get("text") or msg.get("MESSAGE", ""))
        if text:
            lines.append(f"[{date}] **ID{author}:** {text[:800]}")
        for fl in _fmt_msg_files(msg):
            lines.append(fl)
    return "\n".join(lines)

# ─── TRANSFORM: файлы диска ───────────────────────────────────────────────────

def build_files_doc(files: list, group_name: str, inline_sections: list = None) -> str:
    lines = [
        f"# {group_name} — Файлы на диске",
        f"_Сгенерировано: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}_",
        f"_Всего файлов: {len(files)}_",
        ""
    ]
    for f in files:
        name    = f.get("NAME", "")
        size    = f.get("SIZE", "")
        created = fmt_date(f.get("CREATE_TIME", ""))
        ext     = f.get("EXTENSION", "").upper()
        lines.append(f"- **{name}** ({ext}, {size} байт, загружен {created})")
    if inline_sections:
        lines.append("")
        lines.append("## Содержимое небольших текстовых файлов")
        lines.extend(inline_sections)
    return "\n".join(lines)

# ─── TRANSFORM: CRM ───────────────────────────────────────────────────────────

def build_crm_companies_doc(companies: list) -> str:
    lines = [
        "# CRM — Компании",
        f"_Сгенерировано: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}_",
        f"_Всего компаний: {len(companies)}_",
        ""
    ]
    for co in companies:
        name = co.get("TITLE") or f"Компания #{co.get('ID')}"
        lines.append(f"## {name}")
        meta = []
        if co.get("INDUSTRY"):
            meta.append(f"**Отрасль:** {co['INDUSTRY']}")
        phones = _format_multifield(co.get("PHONE"))
        if phones:
            meta.append(f"**Телефон:** {phones}")
        emails = _format_multifield(co.get("EMAIL"))
        if emails:
            meta.append(f"**Email:** {emails}")
        web = _format_multifield(co.get("WEB"))
        if web:
            meta.append(f"**Сайт:** {web}")
        if co.get("ADDRESS"):
            meta.append(f"**Адрес:** {co['ADDRESS']}")
        created = fmt_date(co.get("DATE_CREATE", ""))
        if created:
            meta.append(f"**Создана:** {created}")
        if meta:
            lines.append("  ".join(meta))
        comments = clean_bbcode(co.get("COMMENTS") or "")
        if comments:
            lines.append(f"**Заметки:** {comments[:500]}")
        lines.append("")
        lines.append("---")
        lines.append("")
    return "\n".join(lines)


def build_crm_contacts_doc(contacts: list) -> str:
    lines = [
        "# CRM — Контакты",
        f"_Сгенерировано: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}_",
        f"_Всего контактов: {len(contacts)}_",
        ""
    ]
    for ct in contacts:
        parts = [ct.get("LAST_NAME", ""), ct.get("NAME", ""), ct.get("SECOND_NAME", "")]
        full_name = " ".join(p for p in parts if p).strip() or f"Контакт #{ct.get('ID')}"
        lines.append(f"## {full_name}")
        meta = []
        if ct.get("POST"):
            meta.append(f"**Должность:** {ct['POST']}")
        if ct.get("COMPANY_TITLE"):
            meta.append(f"**Компания:** {ct['COMPANY_TITLE']}")
        phones = _format_multifield(ct.get("PHONE"))
        if phones:
            meta.append(f"**Телефон:** {phones}")
        emails = _format_multifield(ct.get("EMAIL"))
        if emails:
            meta.append(f"**Email:** {emails}")
        created = fmt_date(ct.get("DATE_CREATE", ""))
        if created:
            meta.append(f"**Создан:** {created}")
        if meta:
            lines.append("  ".join(meta))
        comments = clean_bbcode(ct.get("COMMENTS") or "")
        if comments:
            lines.append(f"**Заметки:** {comments[:500]}")
        lines.append("")
        lines.append("---")
        lines.append("")
    return "\n".join(lines)


def build_crm_deals_doc(deals: list) -> str:
    lines = [
        "# CRM — Сделки",
        f"_Сгенерировано: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}_",
        f"_Всего сделок: {len(deals)}_",
        ""
    ]
    for d in deals:
        title = d.get("TITLE") or f"Сделка #{d.get('ID')}"
        lines.append(f"## {title}")
        meta = []
        if d.get("STAGE_ID"):
            meta.append(f"**Стадия:** {d['STAGE_ID']}")
        if d.get("OPPORTUNITY"):
            currency = d.get("CURRENCY_ID", "")
            meta.append(f"**Сумма:** {d['OPPORTUNITY']} {currency}".strip())
        if d.get("CLOSEDATE"):
            meta.append(f"**Дата закрытия:** {fmt_date(d['CLOSEDATE'])}")
        created = fmt_date(d.get("DATE_CREATE", ""))
        if created:
            meta.append(f"**Создана:** {created}")
        if meta:
            lines.append("  ".join(meta))
        comments = clean_bbcode(d.get("COMMENTS") or "")
        if comments:
            lines.append(f"**Комментарии:** {comments[:500]}")
        lines.append("")
        lines.append("---")
        lines.append("")
    return "\n".join(lines)


def build_crm_smart_doc(items: list) -> str:
    lines = [
        "# CRM — Смарт-процессы",
        f"_Сгенерировано: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}_",
        f"_Всего элементов: {len(items)}_",
        ""
    ]
    # Группируем по типу
    from collections import defaultdict
    by_type: dict = defaultdict(list)
    for item in items:
        by_type[item.get("_typeName", "Неизвестный тип")].append(item)
    for type_name, type_items in by_type.items():
        lines.append(f"## {type_name} ({len(type_items)} эл.)")
        lines.append("")
        for it in type_items:
            title = it.get("title") or f"#{it.get('id', '')}"
            stage = it.get("stageId", "")
            created = fmt_date(it.get("createdTime", ""))
            row = f"- **{title}**"
            if stage:
                row += f" | стадия: {stage}"
            if created:
                row += f" | создан: {created}"
            lines.append(row)
        lines.append("")
    return "\n".join(lines)


_ACTIVITY_DIRECTION_LABELS = {"1": "входящая", "2": "исходящая"}


def build_crm_timeline_doc(comments: list, activities: list) -> str:
    """История активности CRM: комментарии таймлайна + завершённые активности,
    сгруппированные по карточке (entity_name)."""
    lines = [
        "# CRM — История активности",
        f"_Сгенерировано: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}_",
        f"_Комментариев: {len(comments)}, активностей: {len(activities)}_",
        ""
    ]
    from collections import defaultdict
    by_entity: dict = defaultdict(lambda: {"comments": [], "activities": []})
    for c in comments:
        by_entity[c.get("_entity_name", "?")]["comments"].append(c)
    for a in activities:
        by_entity[a.get("_entity_name", "?")]["activities"].append(a)

    for entity_name, data in by_entity.items():
        lines.append(f"## {entity_name}")
        if data["comments"]:
            lines.append("**Комментарии:**")
            for cm in data["comments"]:
                dt = fmt_datetime(cm.get("CREATED", ""))
                text = clean_bbcode(cm.get("COMMENT", ""))
                if text:
                    lines.append(f"- [{dt}] {text[:600]}")
        if data["activities"]:
            lines.append("**Активности:**")
            for a in data["activities"]:
                dt = fmt_datetime(a.get("CREATED", ""))
                subj = a.get("SUBJECT", "") or "(без темы)"
                direction = _ACTIVITY_DIRECTION_LABELS.get(str(a.get("DIRECTION", "")), "")
                row = f"- [{dt}] {subj}"
                if direction:
                    row += f" ({direction})"
                lines.append(row)
        lines.append("")
        lines.append("---")
        lines.append("")
    return "\n".join(lines)


# ─── EXTRACT: открытые линии ──────────────────────────────────────────────────

def _get_chat_messages(c: B24Client, chat_id) -> list:
    """Загрузить все сообщения чата по CHAT_ID с пагинацией."""
    messages = []
    last_id = None
    while True:
        params = {"DIALOG_ID": f"chat{chat_id}", "LIMIT": 50}
        if last_id:
            params["LAST_ID"] = last_id
        try:
            raw = c.call("im.dialog.messages.get", params)
        except B24Error:
            break
        if not raw:
            break
        if isinstance(raw, dict):
            batch = raw.get("messages", [])
            if isinstance(batch, dict):
                batch = list(batch.values())
        elif isinstance(raw, list):
            batch = raw
        else:
            break
        if not batch:
            break
        messages.extend(batch)
        # Если пришло меньше лимита — достигли конца
        if len(batch) < 50:
            break
        # Берём минимальный ID для следующей страницы (сообщения идут от новых к старым)
        ids = [m.get("id") or m.get("ID") for m in batch if m.get("id") or m.get("ID")]
        if not ids:
            break
        last_id = min(int(i) for i in ids)
    return messages


def build_specific_chat_doc(messages: list, chat_id: int, group_name: str) -> str:
    """Документ для конкретного чата по его ID."""
    lines = [
        f"# {group_name} — Чат {chat_id}",
        f"_Сгенерировано: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}_",
        f"_Сообщений: {len(messages)}_",
        ""
    ]
    for msg in reversed(messages):
        author = (msg.get("authorName") or msg.get("AUTHOR_NAME")
                  or f"ID{msg.get('authorId', msg.get('AUTHOR_ID', ''))}")
        date   = fmt_datetime(msg.get("date") or msg.get("DATE_CREATE", ""))
        text   = clean_bbcode(msg.get("text") or msg.get("MESSAGE", ""))
        if text:
            lines.append(f"[{date}] **{author}:** {text[:600]}")
        for fl in _fmt_msg_files(msg):
            lines.append(fl)
    return "\n".join(lines)


# ─── SAVE ─────────────────────────────────────────────────────────────────────

def save_docs(docs: dict, out_dir: str):
    Path(out_dir).mkdir(parents=True, exist_ok=True)
    for name, content in docs.items():
        path = Path(out_dir) / name
        path.write_text(content, encoding="utf-8")
        kb = len(content.encode("utf-8")) // 1024
        print(f"  Записан: {path} ({len(content.splitlines())} строк, {kb} KB)")

# ─── helpers ──────────────────────────────────────────────────────────────────

def _get_group_name(c: B24Client, group_id: int) -> str:
    try:
        st = c.call("disk.storage.getlist", {"filter": {"ENTITY_TYPE": "group", "ENTITY_ID": group_id}})
        if st:
            return st[0].get("NAME", f"Проект {group_id}")
    except B24Error:
        pass
    return f"Проект {group_id}"


def _redact_and_save(docs: dict, out_dir: str, prefix: str = "") -> tuple:
    """Редактировать секреты и сохранить. Возвращает (redacted_docs, secrets_total)."""
    print(f"\n{'─'*50}")
    print("🔐 Сканирую документы на секреты...")
    redacted = {}
    total = 0
    for name, content in docs.items():
        label = f"{prefix}{name}" if prefix else name
        cleaned, count = security.redact_doc(content, source_label=label)
        redacted[name] = cleaned
        marker = f"⚠️  найдено {count}" if count else "✅ чисто"
        print(f"  {name}: {marker}")
        total += count
    save_docs(redacted, out_dir)
    return redacted, total


def _load_to_nlm(c_cfg: dict, out_dir: str, doc_names: list,
                 group_name: str, cfg: dict, cfg_path: str, cfg_key: str):
    """Загрузить документы в NotebookLM, сохранить notebook_id обратно в config."""
    import load_notebooklm
    nb_id = c_cfg.get("notebook_id", "")
    new_id = load_notebooklm.load_to_notebooklm(out_dir, nb_id, group_name, doc_names)
    if new_id and new_id != nb_id:
        # Записать новый id обратно в нужное место конфига
        _set_nested(cfg, cfg_key, new_id)
        with open(cfg_path, "w", encoding="utf-8") as fh:
            json.dump(cfg, fh, ensure_ascii=False, indent=2)
        security.harden_config_file(cfg_path)
    print(f"✅ NotebookLM: https://notebooklm.google.com/notebook/{new_id}")
    return new_id


def _set_nested(d: dict, path: str, value):
    """Установить значение по пути 'a.b.c' в словаре."""
    keys = path.split(".")
    for k in keys[:-1]:
        d = d.setdefault(k, {})
    d[keys[-1]] = value


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--config", default="config.json")
    parser.add_argument("--config-url",     default="", help="URL /api/config.php на хостинге")
    parser.add_argument("--status-url",     default="", help="URL /api/status.php на хостинге")
    parser.add_argument("--connector-token", default="", help="X-Connector-Token (или задать CONNECTOR_TOKEN=...)")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--skip-admin-check", action="store_true")
    parser.add_argument("--bind", action="store_true")
    parser.add_argument("--poll", action="store_true")
    parser.add_argument("--group", type=int, help="Обработать только указанный group_id")
    parser.add_argument("--crm-only", action="store_true",
                         help="Пропустить проекты, синхронизировать только глобальный CRM-блок (для matrix-параллелизации)")
    parser.add_argument("--full-inventory", action="store_true",
                         help="Полная инвентаризация аккаунта NotebookLM (list_notebooks) — "
                              "ТОЛЬКО по ручному запросу из админки, не для расписания/webhook")
    args = parser.parse_args()

    connector_token = args.connector_token or os.environ.get("CONNECTOR_TOKEN", "")

    # ── Полная инвентаризация NotebookLM: не трогает Bitrix24/проекты/CRM ────
    if args.full_inventory:
        print("→ Полная инвентаризация NotebookLM (весь аккаунт)...")
        import load_notebooklm
        notebooks = load_notebooklm.list_notebooks(with_source_counts=True)
        payload = {
            "ts": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S"),
            "notebooks_full": notebooks,
        }
        if args.status_url and connector_token:
            post_run_status(args.status_url, connector_token, payload)
            print(f"✅ Инвентаризация отправлена на {args.status_url}: {len(notebooks)} ноутбуков")
        else:
            print(json.dumps(payload, ensure_ascii=False, indent=2))
        return

    # ── Загрузка конфига (с хостинга или из файла) ────────────────────────────
    _run_start = time.time()
    if args.config_url and connector_token:
        print(f"Загружаю конфиг с {args.config_url} ...")
        try:
            cfg = fetch_remote_config(args.config_url, connector_token)
            print("  Конфиг получен с хостинга")
        except Exception as e:
            print(f"  ⚠️  Не удалось получить конфиг с хостинга ({e}), использую локальный")
            cfg = load_config(args.config)
    else:
        cfg     = load_config(args.config)
    webhook = cfg.get("webhook_url") or os.environ.get("BITRIX24_WEBHOOK", "")
    inc     = cfg.get("include", {})
    out_dir = cfg.get("output_dir", "out")
    nlm_on  = cfg.get("targets", {}).get("notebooklm", {}).get("enabled", False)
    dedupe_mode = cfg.get("dedupe_mode", "dry_run")  # "dry_run" | "delete" | "off"

    # Поддержка нового (projects[]) и старого (group_id) форматов конфига
    projects = cfg.get("projects")
    if not projects:
        old_nb = cfg.get("targets", {}).get("notebooklm", {}).get("notebook_id", "")
        projects = [{"group_id": int(cfg["group_id"]), "notebook_id": old_nb}]

    # ── Валидация ─────────────────────────────────────────────────────────────
    cfg_warnings = security.validate_config({**cfg, "webhook_url": webhook})
    for w in cfg_warnings:
        print(f"⚠️  {w}")
    if any(w.startswith("CRITICAL") for w in cfg_warnings):
        sys.exit(1)

    any_chats = any(p.get("chats") for p in projects)
    if nlm_on and any_chats:
        print("⚠️  ВНИМАНИЕ: чаты уходят в NotebookLM (Google, США).\n"
              "   Только для внутреннего использования АНИТ (152-ФЗ ст. 12).")

    security.harden_config_file(args.config)
    security.audit_run_start(cfg, "dry-run" if args.dry_run else "full")

    c = B24Client(webhook, timeout=180)
    print("Подключение к Bitrix24...")
    me = c.call("profile")
    print(f"Авторизован как: {me.get('NAME')} {me.get('LAST_NAME')}")

    if not args.skip_admin_check:
        try:
            security.require_admin(c, strict=True)
            print("✅ Права администратора подтверждены")
        except PermissionError as e:
            print(f"❌ {e}"); sys.exit(1)

    if args.bind:
        print("⚠️  event.bind недоступен для вебхуков. Используйте --poll.")
        return

    # ── --poll: проверить изменения по дате ───────────────────────────────────
    if args.poll:
        state_file = Path(cfg.get("state_file", "state/state.json"))
        last_run = None
        if state_file.exists():
            try:
                last_run = json.loads(state_file.read_text()).get("last_run")
            except Exception:
                pass
        if last_run:
            print(f"\n→ Проверка изменений с {last_run}")
            any_changed = False

            # Задачи
            for p in projects:
                gid = int(p["group_id"])
                try:
                    ch = list(c.call_list("tasks.task.list", {
                        "filter": {"GROUP_ID": gid, ">CHANGED_DATE": last_run},
                        "select": ["ID"],
                    }, result_key="tasks"))
                    if ch:
                        print(f"  Проект {gid} — задачи: {len(ch)} изменений")
                        any_changed = True
                except B24Error as e:
                    print(f"  Проект {gid}: ошибка проверки ({e}) → пересборка")
                    any_changed = True

            # CRM компании и контакты
            crm_chk = inc.get("crm", {})
            if crm_chk.get("enabled"):
                for crm_method, label in [("crm.company.list", "компании"),
                                           ("crm.contact.list", "контакты")]:
                    try:
                        ch = list(c.call_list(crm_method, {
                            "filter": {">DATE_MODIFY": last_run},
                            "select": ["ID"],
                        }))
                        if ch:
                            print(f"  CRM {label}: {len(ch)} изменений")
                            any_changed = True
                    except B24Error:
                        any_changed = True

            # Чаты — IM не имеет «изменён с даты X», пересобираем всегда
            if any_chats:
                print(f"  Чаты: всегда пересобираем (IM не поддерживает инкремент)")
                any_changed = True

            if not any_changed:
                print("  Изменений нет. Пересборка не нужна.")
                return
            print("  → полная пересборка")
        else:
            print("\n→ Первый запуск --poll: полная пересборка")

    secrets_grand_total = 0
    project_errors = []

    # ── Проекты ───────────────────────────────────────────────────────────────
    for idx, proj in enumerate(projects):
        if args.crm_only:
            continue
        if args.group and int(proj["group_id"]) != args.group:
            continue
        try:
            group_id   = int(proj["group_id"])
            group_name = _get_group_name(c, group_id)
            # Заголовок ноутбука: явное название из админ-панели имеет приоритет
            # над именем группы Б24; источники — per-project с фолбэком на include.
            nb_title = (proj.get("notebook_name") or "").strip() or group_name
            src      = proj.get("sources") or {}
            def _want(key):
                return src.get(key, inc.get(key, True))
            proj_dir   = str(Path(out_dir) / f"project_{group_id}")
            print(f"\n{'═'*50}")
            print(f"ПРОЕКТ: {group_name} (group_id={group_id})")

            docs = {}
            chat_messages_for_audio = []
            disk_binary_attachments = []

            if _want("tasks"):
                print("\n→ Задачи")
                tasks = extract_tasks_full(c, group_id)
                comments = extract_comments_bulk(c, [t["id"] for t in tasks]) if _want("task_comments") else None
                docs["tasks.md"] = build_tasks_doc(tasks, group_name, c, comments)

            if _want("group_chat"):
                print("\n→ Чат группы")
                dialog, messages = extract_group_chat(c, group_id)
                if dialog or messages:
                    docs["chat.md"] = build_chat_doc(dialog, messages, group_name)
                    chat_messages_for_audio = messages

            if _want("disk_files"):
                print("\n→ Файлы диска")
                files = extract_disk_files(c, group_id)
                if files:
                    disk_limit = int(proj.get("disk_files_limit", cfg.get("disk_files_limit", 15)))
                    text_docs, disk_binary_attachments, inline_sections = prepare_disk_attachments(
                        c, files, group_name, proj_dir, limit=disk_limit)
                    docs.update(text_docs)
                    docs["files.md"] = build_files_doc(files, group_name, inline_sections)

            _raw_chats = list(proj.get("chats") or [])
            _ol_ids = [str(ol["id"]) for ol in (proj.get("open_lines_data") or [])
                       if isinstance(ol, dict) and ol.get("id")]
            _seen = set(str(x) for x in _raw_chats)
            for chat_id in _raw_chats + [x for x in _ol_ids if x not in _seen]:
                print(f"\n→ Чат {chat_id}")
                try:
                    messages = _get_chat_messages(c, chat_id)
                    print(f"    Сообщений: {len(messages)}")
                    if messages:
                        docs[f"chat_{chat_id}.md"] = build_specific_chat_doc(messages, chat_id, group_name)
                        chat_messages_for_audio.extend(messages)
                except B24Error as e:
                    print(f"  ❌ Ошибка чата {chat_id}: {e}")

            # ── Per-project CRM (точечно по ID компаний) ──────────────────────────
            # Новый формат: src["crm"] = {company_ids, contacts, deals, smart}.
            # Старый формат (crm_companies/crm_contacts/... = «выгрузить всё») больше
            # не поддерживается — CRM выгружается только если заданы company_ids.
            crm_src = src.get("crm")
            proj_company_ids = []
            if isinstance(crm_src, dict):
                proj_company_ids = [str(x) for x in (crm_src.get("company_ids") or []) if x]

            if proj_company_ids:
                print(f"\n→ CRM по компаниям: {', '.join(proj_company_ids)}")
                companies = extract_crm_companies(c, proj_company_ids)
                if companies:
                    docs["crm_companies.md"] = build_crm_companies_doc(companies)

                if crm_src.get("contacts"):
                    contacts = extract_crm_contacts_by_company(c, proj_company_ids)
                    if contacts:
                        docs["crm_contacts.md"] = build_crm_contacts_doc(contacts)

                if crm_src.get("deals"):
                    deals = extract_crm_deals(c, proj_company_ids)
                    if deals:
                        docs["crm_deals.md"] = build_crm_deals_doc(deals)

                if crm_src.get("smart"):
                    smart = extract_crm_smart(c, proj_company_ids)
                    if smart:
                        docs["crm_smart.md"] = build_crm_smart_doc(smart)

                if crm_src.get("timeline"):
                    company_names = {str(co["ID"]): (co.get("TITLE") or f"Компания #{co['ID']}")
                                      for co in companies} if companies else {}
                    entities = [{"type": "company", "id": cid,
                                 "name": company_names.get(str(cid), f"Компания #{cid}")}
                                for cid in proj_company_ids]
                    timeline_comments = extract_crm_timeline(c, entities)
                    timeline_activities = extract_crm_activities(c, entities)
                    if timeline_comments or timeline_activities:
                        docs["crm_timeline.md"] = build_crm_timeline_doc(
                            timeline_comments, timeline_activities)

            redacted, n_secrets = _redact_and_save(docs, proj_dir)
            secrets_grand_total += n_secrets
            security.audit_run_end(redacted, n_secrets)

            if not args.dry_run and nlm_on:
                print(f"\n→ Загрузка в NotebookLM [{nb_title}]")
                try:
                    import load_notebooklm
                    nb_id = proj.get("notebook_id", "")
                    new_id = load_notebooklm.load_to_notebooklm(
                        proj_dir, nb_id, nb_title, list(redacted.keys()),
                        dedupe_mode=dedupe_mode)
                    if new_id and new_id != nb_id:
                        cfg["projects"][idx]["notebook_id"] = new_id
                        if args.config_url and connector_token:
                            try:
                                push_notebook_id_patch(args.config_url, connector_token,
                                                        "project", new_id, group_id=group_id)
                                print("  Конфиг обновлён на хостинге")
                            except Exception as e:
                                print(f"  ⚠️  Не удалось сохранить конфиг на хостинге ({e})")
                        else:
                            save_config(cfg, args, connector_token)
                    print(f"✅ NotebookLM: https://notebooklm.google.com/notebook/{new_id}")
                    # Аудио из чата группы
                    if chat_messages_for_audio:
                        print("  → Аудиофайлы из чата группы")
                        audio_dir = str(Path(proj_dir) / "audio")
                        audio_list = download_audio_from_messages(chat_messages_for_audio, audio_dir)
                        if audio_list:
                            load_notebooklm.upload_files(new_id, audio_list)
                    if disk_binary_attachments:
                        print("  → Файлы диска as-is (PDF/изображения)")
                        load_notebooklm.upload_files(new_id, disk_binary_attachments)
                except Exception as e:
                    print(f"❌ NotebookLM: {e}")
        except Exception as e:
            gid = int(proj.get("group_id", 0))
            gname = (proj.get("notebook_name") or "").strip() or f"group {gid}"
            print(f"\n❌ ПРОЕКТ {gname} (group_id={gid}): {e}")
            project_errors.append({"group_id": gid, "name": gname, "error": str(e)})
            continue

    # ── CRM (только при полном прогоне, не при --group) ───────────────────────
    crm_cfg = inc.get("crm", {})
    if args.group:
        crm_cfg = {}  # пропускаем CRM при синхронизации одного проекта
    if crm_cfg.get("enabled"):
        crm_dir = str(Path(out_dir) / "crm")
        crm_docs = {}
        print(f"\n{'═'*50}")
        print("CRM ДАННЫЕ")
        company_ids = crm_cfg.get("company_ids", [])
        contact_ids = crm_cfg.get("contact_ids", [])
        if company_ids:
            print("\n→ Компании")
            companies = extract_crm_companies(c, company_ids)
            if companies:
                crm_docs["crm_companies.md"] = build_crm_companies_doc(companies)
        if contact_ids:
            print("\n→ Контакты")
            contacts = extract_crm_contacts(c, contact_ids)
            if contacts:
                crm_docs["crm_contacts.md"] = build_crm_contacts_doc(contacts)
        if not company_ids and not contact_ids:
            print("  (нет выбранных сущностей — настройте company_ids/contact_ids)")

        if crm_cfg.get("timeline") and (company_ids or contact_ids):
            print("\n→ История активности (timeline)")
            company_names = {str(co["ID"]): (co.get("TITLE") or f"Компания #{co['ID']}")
                              for co in (companies if company_ids else [])}
            contact_names = {}
            for ct in (contacts if contact_ids else []):
                full = " ".join(p for p in [ct.get("LAST_NAME", ""), ct.get("NAME", "")] if p).strip()
                contact_names[str(ct["ID"])] = full or f"Контакт #{ct['ID']}"
            entities = (
                [{"type": "company", "id": cid, "name": company_names.get(str(cid), f"Компания #{cid}")}
                 for cid in company_ids] +
                [{"type": "contact", "id": cid, "name": contact_names.get(str(cid), f"Контакт #{cid}")}
                 for cid in contact_ids]
            )
            timeline_comments = extract_crm_timeline(c, entities)
            timeline_activities = extract_crm_activities(c, entities)
            if timeline_comments or timeline_activities:
                crm_docs["crm_timeline.md"] = build_crm_timeline_doc(timeline_comments, timeline_activities)
        if crm_docs:
            redacted_crm, n_sec = _redact_and_save(crm_docs, crm_dir)
            secrets_grand_total += n_sec
            security.audit_run_end(redacted_crm, n_sec)
            if not args.dry_run and nlm_on:
                print("\n→ Загрузка CRM в NotebookLM")
                try:
                    import load_notebooklm
                    nb_id = crm_cfg.get("notebook_id", "")
                    new_id = load_notebooklm.load_to_notebooklm(
                        crm_dir, nb_id, "CRM АНИТ", list(redacted_crm.keys()),
                        dedupe_mode=dedupe_mode)
                    if new_id and new_id != nb_id:
                        cfg["include"]["crm"]["notebook_id"] = new_id
                        if args.config_url and connector_token:
                            try:
                                push_notebook_id_patch(args.config_url, connector_token,
                                                        "crm", new_id)
                                print("  Конфиг обновлён на хостинге")
                            except Exception as e:
                                print(f"  ⚠️  Не удалось сохранить конфиг на хостинге ({e})")
                        else:
                            save_config(cfg, args, connector_token)
                    print(f"✅ NotebookLM CRM: https://notebooklm.google.com/notebook/{new_id}")
                except Exception as e:
                    print(f"❌ NotebookLM CRM: {e}")

    # ── Сохраняем last_run ────────────────────────────────────────────────────
    now_iso = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S")
    state_file = Path(cfg.get("state_file", "state/state.json"))
    state_file.parent.mkdir(parents=True, exist_ok=True)
    state_file.write_text(json.dumps({"last_run": now_iso}, ensure_ascii=False))

    elapsed = round(time.time() - _run_start, 1)

    # ── POST статуса на хостинг ───────────────────────────────────────────────
    if args.status_url and connector_token:
        payload = {
            "ts":             now_iso,
            "mode":           "dry-run" if args.dry_run else ("poll" if args.poll else "full"),
            "status":         "partial" if project_errors else "ok",
            "secrets_found":  secrets_grand_total,
            "elapsed_sec":    elapsed,
            "group":          args.group,
            "partial_errors": project_errors,
        }
        # Снимок NotebookLM: статус сессии + число источников по ИЗВЕСТНЫМ
        # ноутбукам проектов (не весь аккаунт NotebookLM — там могут быть
        # десятки посторонних личных ноутбуков, и опрос каждого из них по
        # отдельности медленно/ненадёжно на ограниченном по времени прогоне).
        if nlm_on:
            try:
                import load_notebooklm
                payload["nlm_session"] = load_notebooklm.session_status()
                # При matrix-параллелизации (--group/--crm-only) каждый job знает
                # только про свой проект из cfg, снятого в начале ЭТОГО прогона —
                # notebook_id, только что сохранённые СОСЕДНИМИ параллельными
                # job'ами, в нём ещё не отражены. Поэтому перечитываем актуальный
                # конфиг с хостинга прямо перед снимком (GET дешёвый и безопасный,
                # в отличие от полной перезаписи).
                snapshot_cfg = cfg
                if args.config_url and connector_token:
                    try:
                        snapshot_cfg = fetch_remote_config(args.config_url, connector_token)
                    except Exception as e:
                        print(f"  ⚠️  Не удалось перечитать конфиг для снимка ({e}), использую локальный")
                known_notebooks = [
                    {"id": p.get("notebook_id"), "name": (p.get("notebook_name") or "").strip()
                                                          or _get_group_name(c, int(p["group_id"]))}
                    for p in snapshot_cfg.get("projects", []) if p.get("notebook_id")
                ]
                crm_nb_id = snapshot_cfg.get("include", {}).get("crm", {}).get("notebook_id")
                if crm_nb_id:
                    known_notebooks.append({"id": crm_nb_id, "name": "CRM АНИТ"})
                payload["notebooks"] = load_notebooklm.project_notebooks_snapshot(known_notebooks)
            except Exception as e:
                print(f"  ⚠️  Снимок NotebookLM не собран: {e}")
        print(f"\n→ Отправляю статус на {args.status_url} ...")
        post_run_status(args.status_url, connector_token, payload)

    if args.dry_run:
        print(f"\n[dry-run] Готово. Секретов замаскировано: {secrets_grand_total}. Время: {elapsed}s")
        return

    print(f"\n{'═'*50}")
    print(f"✅ Всё готово. Секретов замаскировано: {secrets_grand_total}. Время: {elapsed}s")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        sys.exit(0)
    except Exception as _exc:
        # Попытка отправить статус ошибки на хостинг, если --status-url задан
        import traceback
        traceback.print_exc()
        import sys as _sys
        _args = _sys.argv
        _status_url = ""
        _token = ""
        for _i, _a in enumerate(_args):
            if _a == "--status-url" and _i + 1 < len(_args):
                _status_url = _args[_i + 1]
            if _a == "--connector-token" and _i + 1 < len(_args):
                _token = _args[_i + 1]
        _token = _token or os.environ.get("CONNECTOR_TOKEN", "")
        if _status_url and _token:
            post_run_status(_status_url, _token, {
                "ts":     datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S"),
                "mode":   "unknown",
                "status": "error",
                "error":  str(_exc),
            })
        sys.exit(1)
